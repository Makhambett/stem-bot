import os
from datetime import datetime
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery
from openpyxl import load_workbook
from bot.config import MANAGER_CHAT_ID
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup


router = Router()

EXCEL_FILE = "orders.xlsx"


def get_stats_text():
    if not os.path.exists(EXCEL_FILE):
        return "📭 Заявок пока нет."

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        return "📭 Заявок пока нет."

    total = len(rows)
    today = datetime.now().strftime("%d.%m.%Y")
    today_count = sum(1 for r in rows if r[0] and r[0].startswith(today))

    categories = [r[3] for r in rows if r[3]]
    top_category = max(set(categories), key=categories.count) if categories else "—"
    top_count = categories.count(top_category)

    last_5 = rows[-5:][::-1]
    last_text = "\n".join(
        f"• {r[1]} — {r[3]} — {r[0][:11] if r[0] else '—'}"
        for r in last_5
    )

    return (
        f"📊 <b>Статистика заявок</b>\n\n"
        f"📋 Всего заявок: {total}\n"
        f"📅 Сегодня: {today_count}\n"
        f"🏆 Топ категория: {top_category} ({top_count} заявок)\n\n"
        f"🕐 <b>Последние заявки:</b>\n{last_text}"
    )


@router.message(Command("stats"))
async def show_stats(message: Message):
    if message.from_user.id != MANAGER_CHAT_ID:
        return
    await message.answer(get_stats_text(), parse_mode="HTML")


@router.message(Command("manager"))
async def manager_panel(message: Message):
    if message.from_user.id != MANAGER_CHAT_ID:
        return
    from bot.keyboards.main_menu import manager_menu_kb
    await message.answer(
        "👨‍💼 <b>Панель менеджера</b>",
        parse_mode="HTML",
        reply_markup=manager_menu_kb()
    )


@router.callback_query(F.data == "manager_stats")
async def stats_callback(call: CallbackQuery):
    if call.from_user.id != MANAGER_CHAT_ID:
        return
    await call.message.edit_text(get_stats_text(), parse_mode="HTML")


@router.callback_query(F.data == "manager_orders")
async def last_orders_callback(call: CallbackQuery):
    if call.from_user.id != MANAGER_CHAT_ID:
        return

    if not os.path.exists(EXCEL_FILE):
        await call.message.edit_text("📭 Заявок пока нет.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        await call.message.edit_text("📭 Заявок пока нет.")
        return

    last_10 = rows[-10:][::-1]
    text = "📋 <b>Последние заявки:</b>\n\n" + "\n".join(
        f"👤 {r[1]} | 📞 {r[2]} | 📦 {r[3]} | {r[0][:11] if r[0] else '—'}"
        for r in last_10
    )
    await call.message.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data == "open_catalog")
async def open_catalog_callback(call: CallbackQuery):
    if call.from_user.id != MANAGER_CHAT_ID:
        return
    from bot.keyboards.main_menu import main_menu_kb
    await call.message.edit_text(
        "📦 <b>Каталог</b>\n\nВыберите категорию:",
        parse_mode="HTML",
        reply_markup=main_menu_kb()
    )

class SearchForm(StatesGroup):
    query = State()


@router.callback_query(F.data == "manager_search")
async def search_start(call: CallbackQuery, state: FSMContext):
    if call.from_user.id != MANAGER_CHAT_ID:
        return
    await call.message.edit_text("🔍 Введите имя или телефон для поиска:")
    await state.set_state(SearchForm.query)


@router.message(SearchForm.query)
async def search_orders(message: Message, state: FSMContext):
    if message.from_user.id != MANAGER_CHAT_ID:
        return

    query = message.text.lower()

    if not os.path.exists(EXCEL_FILE):
        await message.answer("📭 Заявок пока нет.")
        await state.clear()
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    results = [
        r for r in rows
        if (r[1] and query in str(r[1]).lower()) or
           (r[2] and query in str(r[2]).lower())
    ]

    if not results:
        await message.answer("❌ Ничего не найдено.")
        await state.clear()
        return

    text = f"🔍 <b>Найдено {len(results)} заявок:</b>\n\n" + "\n".join(
        f"👤 {r[1]} | 📞 {r[2]} | 📦 {r[3]} | {r[0][:11] if r[0] else '—'}"
        for r in results
    )

    from bot.keyboards.main_menu import manager_menu_kb
    await message.answer(text, parse_mode="HTML", reply_markup=manager_menu_kb())
    await state.clear()

@router.callback_query(F.data == "manager_clients")
async def clients_list(call: CallbackQuery):
    if call.from_user.id != MANAGER_CHAT_ID:
        return

    if not os.path.exists(EXCEL_FILE):
        await call.message.edit_text("📭 Заявок пока нет.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        await call.message.edit_text("📭 Заявок пока нет.")
        return

    # Уникальные клиенты по Telegram ID
    seen = set()
    clients = []
    for r in rows:
        client_id = r[5]  # ID колонка
        if client_id not in seen:
            seen.add(client_id)
            clients.append(r)

    text = f"👥 <b>Список клиентов ({len(clients)}):</b>\n\n" + "\n".join(
        f"👤 {r[1]} | {r[4] or '—'} | 📞 {r[2]} | первая заявка: {r[0][:11] if r[0] else '—'}"
        for r in clients
    )

    from bot.keyboards.main_menu import manager_menu_kb
    await call.message.edit_text(text, parse_mode="HTML", reply_markup=manager_menu_kb())
