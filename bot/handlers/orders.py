import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import CallbackQuery, Message, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.exceptions import TelegramBadRequest
from bot.config import MANAGER_CHAT_ID
from bot.keyboards.main_menu import main_menu_kb, manager_menu_kb
from bot.cart import get_cart, clear_cart


router = Router()

EXCEL_FILE = "orders.xlsx"

# ─── СЛОВАРЬ АКТИВНЫХ ДИАЛОГОВ ────────────────────────
# { client_id: {"name": "Ерасыл", "username": "@user"} }
active_dialogs: dict = {}


def save_to_excel(name, phone, category, username, user_id):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"
        ws.append(["Дата", "Имя", "Телефон", "Категория", "Telegram", "ID"])
    ws.append([
        datetime.now().strftime("%d.%m.%Y %H:%M"),
        name, phone, category,
        f"@{username}" if username else "—",
        user_id
    ])
    wb.save(EXCEL_FILE)


def is_valid_phone(phone: str) -> bool:
    cleaned = re.sub(r'[\s\-\(\)]', '', phone)
    return bool(re.match(r'^(\+7|8|7)\d{10}$', cleaned))


class OrderForm(StatesGroup):
    name = State()
    phone = State()


class ReplyForm(StatesGroup):
    message_text = State()


class QuestionForm(StatesGroup):
    waiting = State()


MANAGER_REPLY_BUTTONS = {
    "📊 Статистика":       "manager_stats",
    "📋 Последние заявки": "manager_orders",
    "🔍 Поиск заявки":     "manager_search",
    "👥 Список клиентов":  "manager_clients",
    "🛍 Каталог":          "open_catalog",
}


@router.message(F.text.in_(MANAGER_REPLY_BUTTONS.keys()))
async def handle_manager_reply_buttons(message: Message):
    if message.from_user.id != MANAGER_CHAT_ID:
        return
    callback_data = MANAGER_REPLY_BUTTONS[message.text]
    await message.answer(
        "Выберите действие:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=message.text, callback_data=callback_data)]
        ])
    )


# ─── КОРЗИНА ───────────────────────────────────────────

@router.callback_query(F.data == "show_cart")
async def show_cart(call: CallbackQuery):
    cart = get_cart(call.from_user.id)
    await call.answer()

    if not cart:
        await call.message.answer("🧺 Корзина пуста!\n\nДобавьте товары из каталога.")
        return

    text = "🧺 <b>Ваша корзина:</b>\n\n"
    for i, item in enumerate(cart, 1):
        text += f"{i}. {item['name']}\n   📐 {item['sizes']} | 🔖 {item['article']}\n\n"
    text += f"<b>Итого товаров: {len(cart)}</b>"

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Оформить заявку",  callback_data="cart_order")],
        [InlineKeyboardButton(text="🗑 Очистить корзину", callback_data="cart_clear")],
        [InlineKeyboardButton(text="◀️ Назад в меню",    callback_data="main_menu")],
    ])
    await call.message.answer(text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data == "cart_clear")
async def cart_clear(call: CallbackQuery):
    clear_cart(call.from_user.id)
    await call.answer("🗑 Корзина очищена!")
    await call.message.edit_text(
        "🧺 Корзина очищена!\n\nВернитесь в каталог и добавьте товары.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="◀️ Назад в меню", callback_data="main_menu")]
        ])
    )


@router.callback_query(F.data == "cart_order")
async def cart_order_start(call: CallbackQuery, state: FSMContext):
    cart = get_cart(call.from_user.id)
    if not cart:
        await call.answer("Корзина пуста!")
        return
    await call.answer()
    try:
        await call.message.edit_text(
            "📝 <b>Оформление заявки</b>\n\nВведите ваше имя:\n\n"
            "❌ /cancel — отменить заявку",
            parse_mode="HTML"
        )
    except TelegramBadRequest:
        await call.message.delete()
        await call.message.answer(
            "📝 <b>Оформление заявки</b>\n\nВведите ваше имя:\n\n"
            "❌ /cancel — отменить заявку",
            parse_mode="HTML"
        )
    await state.update_data(from_cart=True)
    await state.set_state(OrderForm.name)


@router.callback_query(F.data.startswith("order_"))
async def start_order(call: CallbackQuery, state: FSMContext):
    category = call.data.replace("order_", "")
    await state.update_data(category=category)
    await call.answer()
    try:
        await call.message.edit_text(
            "📝 <b>Оформление заявки</b>\n\nВведите ваше имя:\n\n"
            "❌ /cancel — отменить заявку",
            parse_mode="HTML"
        )
    except Exception:
        await call.message.delete()
        await call.message.answer(
            "📝 <b>Оформление заявки</b>\n\nВведите ваше имя:\n\n"
            "❌ /cancel — отменить заявку",
            parse_mode="HTML"
        )
    await state.set_state(OrderForm.name)


# ─── ОТМЕНА ЗАЯВКИ ────────────────────────────────────

@router.message(Command("cancel"))
async def cancel_order(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state in [OrderForm.name, OrderForm.phone]:
        await state.clear()
        await message.answer(
            "❌ Заявка отменена.\n\nВозвращаемся в меню:",
            reply_markup=main_menu_kb()
        )
    else:
        await message.answer("Нет активной заявки для отмены.")


# ─── ОФОРМЛЕНИЕ ЗАЯВКИ ────────────────────────────────

@router.message(OrderForm.name)
async def get_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer(
        "📞 Введите ваш номер телефона:\n\n"
        "Форматы: +77001234567 или 87001234567\n"
        "❌ /cancel — отменить заявку"
    )
    await state.set_state(OrderForm.phone)


@router.message(OrderForm.phone)
async def get_phone(message: Message, state: FSMContext):
    phone = message.text

    if not is_valid_phone(phone):
        await message.answer(
            "❌ <b>Неверный формат номера!</b>\n\n"
            "Введите номер в формате:\n"
            "• +7 700 123 45 67\n"
            "• 8 700 123 45 67\n"
            "• 87001234567\n\n"
            "❌ /cancel — отменить заявку",
            parse_mode="HTML"
        )
        return

    data = await state.get_data()
    name = data.get("name")
    from_cart = data.get("from_cart", False)

    if from_cart:
        cart = get_cart(message.from_user.id)
        category = "\n".join([f"• {item['name']} | {item['article']}" for item in cart])
        clear_cart(message.from_user.id)
    else:
        category = data.get("category", "—")

    save_to_excel(name, phone, category, message.from_user.username, message.from_user.id)

    # Сохраняем клиента в активные диалоги
    active_dialogs[message.from_user.id] = {
        "name": name,
        "username": message.from_user.username or "—"
    }

    await message.answer(
        f"✅ <b>Заявка принята!</b>\n\n"
        f"Имя: {name}\n"
        f"Телефон: {phone}\n\n"
        f"Менеджер свяжется с вами в ближайшее время.\n"
        f"📞 +7 (700) 088 0132",
        parse_mode="HTML",
        reply_markup=main_menu_kb()
    )

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=f"💬 Ответить {name}",
            callback_data=f"reply_{message.from_user.id}"
        )]
    ])

    await message.bot.send_message(
        chat_id=MANAGER_CHAT_ID,
        text=f"🔔 <b>Новая заявка!</b>\n\n"
             f"👤 Имя: {name}\n"
             f"📞 Телефон: {phone}\n"
             f"📦 Товары:\n{category}\n"
             f"🆔 Telegram: @{message.from_user.username or '—'}\n"
             f"💬 ID: {message.from_user.id}",
        parse_mode="HTML",
        reply_markup=keyboard
    )

    await state.clear()


# ─── ЗАВЕРШЕНИЕ ДИАЛОГА ────────────────────────────────

@router.message(Command("stop"))
async def stop_dialog(message: Message, state: FSMContext):
    await state.clear()
    if message.from_user.id == MANAGER_CHAT_ID:
        # Показываем список активных диалогов если есть
        if active_dialogs:
            buttons = [
                [InlineKeyboardButton(
                    text=f"💬 {info['name']} (@{info['username']})",
                    callback_data=f"reply_{cid}"
                )]
                for cid, info in active_dialogs.items()
            ]
            await message.answer(
                "🔴 Диалог завершён.\n\n"
                "📋 Активные диалоги — выберите кому ответить:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
            )
        else:
            await message.answer("🔴 Диалог завершён.", reply_markup=manager_menu_kb())
    else:
        active_dialogs.pop(message.from_user.id, None)
        await message.answer("🔴 Диалог завершён.", reply_markup=main_menu_kb())


# ─── ОТВЕТ МЕНЕДЖЕРА ──────────────────────────────────

@router.callback_query(F.data.startswith("reply_"))
async def reply_start(call: CallbackQuery, state: FSMContext):
    if call.from_user.id != MANAGER_CHAT_ID:
        return
    client_id = int(call.data.replace("reply_", ""))
    client_info = active_dialogs.get(client_id, {})
    client_name = client_info.get("name", "Клиент")

    await state.update_data(client_id=client_id)
    await call.message.answer(
        f"✍️ Пишете клиенту: <b>{client_name}</b>\n\n"
        f"/stop — завершить диалог",
        parse_mode="HTML"
    )
    await state.set_state(ReplyForm.message_text)
    await call.answer()


@router.message(ReplyForm.message_text, F.text != "/stop")
async def reply_send(message: Message, state: FSMContext):
    if message.from_user.id != MANAGER_CHAT_ID:
        return
    data = await state.get_data()
    client_id = data.get("client_id")
    client_info = active_dialogs.get(client_id, {})
    client_name = client_info.get("name", "Клиент")

    await message.bot.send_message(
        chat_id=client_id,
        text=f"📩 <b>Сообщение от менеджера STEM Academia:</b>\n\n{message.text}",
        parse_mode="HTML"
    )
    await message.answer(
        f"✅ Отправлено → <b>{client_name}</b>\n"
        f"Продолжайте писать или /stop чтобы завершить.",
        parse_mode="HTML"
    )


# ─── ВОПРОС С САЙТА ───────────────────────────────────

@router.message(QuestionForm.waiting, F.text != "/stop")
async def receive_question(message: Message, state: FSMContext):
    question = message.text
    user = message.from_user

    # Добавляем в активные диалоги
    active_dialogs[user.id] = {
        "name": user.full_name,
        "username": user.username or "—"
    }

    await message.answer(
        "✅ Сообщение отправлено менеджеру!\n"
        "Ожидайте ответа в этом чате 💬"
    )

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=f"💬 Ответить {user.full_name}",
            callback_data=f"reply_{user.id}"
        )]
    ])

    await message.bot.send_message(
        chat_id=MANAGER_CHAT_ID,
        text=f"❓ <b>Вопрос с сайта!</b>\n\n"
             f"👤 {user.full_name}\n"
             f"🆔 @{user.username or '—'}\n\n"
             f"💬 {question}",
        parse_mode="HTML",
        reply_markup=keyboard
    )


# ─── REPLY КНОПКИ КЛИЕНТА ─────────────────────────────

REPLY_BUTTONS = {
    "🪑 Мебель":             "cat_mebel",
    "💡 Электротехника":     "cat_electro",
    "🎨 Декор":              "cat_decor",
    "💻 Цифровые продукты":  "cat_digital",
    "🔬 Оборудование":       "cat_equipment",
    "🧺 Корзина":            "show_cart",
    "📞 Контакты":           "faq",
}


@router.message(F.text.in_(REPLY_BUTTONS.keys()))
async def handle_reply_buttons(message: Message, state: FSMContext):
    callback_data = REPLY_BUTTONS[message.text]
    if callback_data == "show_cart":
        cart = get_cart(message.from_user.id)
        if not cart:
            await message.answer("🧺 Корзина пуста!\n\nДобавьте товары из каталога.")
            return
        text = "🧺 <b>Ваша корзина:</b>\n\n"
        for i, item in enumerate(cart, 1):
            text += f"{i}. {item['name']}\n   📐 {item['sizes']} | 🔖 {item['article']}\n\n"
        text += f"<b>Итого товаров: {len(cart)}</b>"
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Оформить заявку",  callback_data="cart_order")],
            [InlineKeyboardButton(text="🗑 Очистить корзину", callback_data="cart_clear")],
        ])
        await message.answer(text, parse_mode="HTML", reply_markup=keyboard)
    else:
        await message.answer(
            "Выберите товар из каталога:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Открыть каталог", callback_data=callback_data)]
            ])
        )
